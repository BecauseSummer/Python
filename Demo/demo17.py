# –*–coding:utf-8 –*–
# 2021-04-30 11:27
from openpyxl import load_workbook

wb = load_workbook('file/1619753319.xlsx')

print(wb.sheetnames)
